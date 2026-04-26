#!/usr/bin/env python3
"""
Excel → PDF → watsonx Text Extraction → JSONL pipeline.

Pipeline steps per input .xlsx/.xlsm/.xls file:
    1. Extract native embedded images from the workbook (xl/media/* inside
       the zip) to <output>/<stem>/images/ — this is the most reliable way
       to capture images, since LibreOffice/PDF rendering sometimes drops
       floating shapes.
    2. Convert to PDF via LibreOffice (preserves formatting / layout).
    3. Upload PDF to COS and run watsonx Text Extraction V2 with
       CREATE_EMBEDDED_IMAGES=enabled_verbatim, so the resulting markdown
       contains real inline images (not just placeholders).
    4. Download all result artifacts (assembly.md, any image files) from
       COS to <output>/<stem>/watsonx/ .
    5. Append one row per processed file to the JSONL output.

Usage (from repo root, venv active):
    source venv/bin/activate

    # Single file
    python excel_extraction/excel_to_jsonl_pipeline.py \\
        "client_data/data/F230103-2208_WindowFit_측면고정_스크류_체결강도_검토_221028.xlsx"

    # A folder (all .xlsx/.xlsm/.xls inside, recursively)
    python excel_extraction/excel_to_jsonl_pipeline.py client_data/data/

    # Custom output directory and JSONL name
    python excel_extraction/excel_to_jsonl_pipeline.py client_data/data/ \\
        --output-dir excel_extraction/output \\
        --jsonl excel_extraction/output/excel_extracted.jsonl

    # Re-use existing PDFs (skip LibreOffice step)
    python excel_extraction/excel_to_jsonl_pipeline.py some_folder --skip-conversion

Requires:
    - LibreOffice (`brew install --cask libreoffice` on macOS)
    - watsonx credentials in .env (WATSONX_APIKEY, WATSONX_URL, SPACE_ID,
      COS_BUCKET_NAME)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import time
import uuid
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Make the watsonx_text_extraction package (sibling) and src/ (for `common.*`)
# importable when running this script directly from the repo root.
_SRC_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_SRC_ROOT))
sys.path.insert(0, str(_SRC_ROOT / "watsonx_text_extraction"))

try:
    from text_extraction import NotebookTextExtraction  # type: ignore[import-not-found]
    from ibm_watsonx_ai.helpers import DataConnection, S3Location
    from ibm_watsonx_ai.metanames import TextExtractionsV2ParametersMetaNames as PMN
    from common.libreoffice import find_libreoffice, convert_to_pdf  # noqa: F401
except ImportError as e:
    print(f"❌ Import error: {e}")
    print("💡 Make sure you're running from the repo root with the venv active.")
    sys.exit(1)


EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}


# --------------------------------------------------------------------------- #
# Input resolution
# --------------------------------------------------------------------------- #

def resolve_inputs(input_path: Path) -> List[Path]:
    """Return a list of .xlsx/.xlsm/.xls files for a file-or-folder argument."""
    input_path = input_path.resolve()
    if input_path.is_file():
        if input_path.suffix.lower() not in EXCEL_EXTS:
            raise ValueError(
                f"Input file is not an Excel workbook: {input_path.name}"
            )
        return [input_path]
    if input_path.is_dir():
        files = sorted(
            p for p in input_path.rglob("*")
            if p.is_file()
            and p.suffix.lower() in EXCEL_EXTS
            and not p.name.startswith("~$")  # skip Excel lock files
            and not p.name.startswith(".")
        )
        return files
    raise FileNotFoundError(f"Path not found: {input_path}")


# --------------------------------------------------------------------------- #
# Image extraction (native, from the .xlsx zip)
# --------------------------------------------------------------------------- #

def extract_images_from_workbook(xlsx_path: Path, images_dir: Path) -> List[str]:
    """Pull embedded images straight out of the .xlsx/.xlsm zip container.

    .xlsx/.xlsm are OOXML zip files; embedded pictures live under xl/media/.
    This is more reliable than rendering through LibreOffice because it
    preserves the original bitmap data and catches images even when the
    workbook's sheet layout wouldn't paginate them cleanly.

    Returns the list of written image paths (as strings).
    """
    ext = xlsx_path.suffix.lower()
    if ext == ".xls":
        # Old binary format — not a zip, can't do this trick. The watsonx
        # extraction step will still catch visible images via OCR.
        return []

    images_dir.mkdir(parents=True, exist_ok=True)
    saved: List[str] = []
    try:
        with zipfile.ZipFile(xlsx_path, "r") as z:
            for member in z.namelist():
                if not member.startswith("xl/media/") or member.endswith("/"):
                    continue
                out_path = images_dir / Path(member).name
                with open(out_path, "wb") as fh:
                    fh.write(z.read(member))
                saved.append(str(out_path))
    except zipfile.BadZipFile:
        print(f"   ⚠️  {xlsx_path.name} isn't a valid zip — can't extract native images")
        return []
    return saved


# --------------------------------------------------------------------------- #
# Per-sheet workbook splitting
# --------------------------------------------------------------------------- #

def _safe_sheet_slug(name: str, idx: int) -> str:
    """Turn a sheet name into a filesystem-safe slug.

    Keeps Korean / CJK characters and alphanumerics, replaces everything else
    (parens, spaces, punctuation) with underscores. Prepends the sheet index
    so the eventual concatenated output stays in sheet order when sorted.
    """
    slug = re.sub(r"[^\w\uac00-\ud7a3\u3040-\u30ff\u4e00-\u9fff]+", "_", name)
    slug = slug.strip("_") or "sheet"
    return f"{idx:02d}_{slug}"


def split_workbook_per_sheet(
    xlsx_path: Path, output_dir: Path
) -> List[Tuple[str, Path]]:
    """Create one .xlsx per sheet in the source workbook.

    Each split file only contains a single worksheet — and, because images /
    drawings are attached to individual worksheets, images belonging to other
    sheets are dropped from its xl/media/. That means the existing
    extract_images_from_workbook(…) applied to a split file gives us
    per-sheet image isolation for free.

    Returns a list of (sheet_name, split_xlsx_path) preserving sheet order.
    .xls (binary) isn't supported; in that case the caller should fall back
    to whole-workbook mode.
    """
    if xlsx_path.suffix.lower() == ".xls":
        raise ValueError(".xls (binary) can't be split; use whole-workbook mode")

    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for per-sheet mode") from exc

    output_dir.mkdir(parents=True, exist_ok=True)

    probe = load_workbook(xlsx_path, read_only=True, data_only=False, keep_vba=False)
    sheet_names = list(probe.sheetnames)
    probe.close()

    out: List[Tuple[str, Path]] = []
    for idx, sheet_name in enumerate(sheet_names, 1):
        wb = load_workbook(xlsx_path, keep_vba=False, data_only=False)
        for other in list(wb.sheetnames):
            if other != sheet_name:
                del wb[other]
        # A workbook must contain at least one *visible* sheet. If the
        # surviving sheet is hidden (sheet_state = "hidden" / "veryHidden"),
        # openpyxl raises "The only worksheet of a workbook cannot be hidden"
        # on save. Force it visible so LibreOffice also renders it.
        only_ws = wb[wb.sheetnames[0]]
        if only_ws.sheet_state != "visible":
            only_ws.sheet_state = "visible"
        split_path = output_dir / f"{_safe_sheet_slug(sheet_name, idx)}.xlsx"
        wb.save(split_path)
        out.append((sheet_name, split_path))
    return out


# --------------------------------------------------------------------------- #
# LibreOffice xlsx → pdf
# --------------------------------------------------------------------------- #

def check_libreoffice() -> Optional[str]:
    """Return a usable LibreOffice binary path, or None."""
    return find_libreoffice()


def convert_excel_to_pdf(
    input_file: Path, output_dir: Path, libreoffice_cmd: Optional[str] = None
) -> Path:
    """Convert a single Excel workbook to PDF via LibreOffice.

    Thin wrapper over `common.libreoffice.convert_to_pdf`. Kept so the rest
    of this pipeline can stay unchanged.
    """
    return convert_to_pdf(
        input_file, output_dir, timeout=180, soffice_cmd=libreoffice_cmd
    )


# --------------------------------------------------------------------------- #
# watsonx extraction
# --------------------------------------------------------------------------- #

# Valid values for CREATE_EMBEDDED_IMAGES on the current watsonx API
# (2026-04-01):
#   disabled                  → images silently dropped
#   enabled_placeholder       → just "[IMAGE]"-style placeholder text (default
#                               in NotebookTextExtraction; why you saw no
#                               image info before)
#   enabled_text              → OCR the text inside each image
#   enabled_verbalization     → vision model writes a caption for images that
#                               likely carry information
#   enabled_verbalization_all → vision model captions every image
_DEFAULT_EMBEDDED_IMAGES = "enabled_verbalization_all"


def _make_parameters_builder(embedded_images_mode: str):
    """Return a zero-arg builder to drop onto a NotebookTextExtraction
    instance, so we don't mutate the upstream default.
    """
    def _build() -> dict:
        return {
            PMN.MODE: "high_quality",
            PMN.OCR_MODE: "enabled",
            PMN.LANGUAGES: ["en", "ko"],
            PMN.AUTO_ROTATION_CORRECTION: True,
            PMN.CREATE_EMBEDDED_IMAGES: embedded_images_mode,
            PMN.OUTPUT_DPI: 200,
            PMN.OUTPUT_TOKENS_AND_BBOX: True,
        }
    return _build


def extract_pdf_with_watsonx(
    extractor: NotebookTextExtraction,
    pdf_path: Path,
    local_output_dir: Path,
    embedded_images_mode: str = _DEFAULT_EMBEDDED_IMAGES,
) -> Optional[Dict]:
    """Upload PDF to COS, run watsonx Text Extraction, download all results
    (markdown + image assets) into local_output_dir. Returns dict with the
    downloaded markdown text and list of downloaded asset paths.
    """
    filename = pdf_path.name
    local_output_dir.mkdir(parents=True, exist_ok=True)

    extractor.cos_client.upload_file(
        str(pdf_path), extractor.cos_bucket_name, filename
    )

    document_reference = DataConnection(
        connection_asset_id=extractor.cos_connection_id,
        location=S3Location(bucket=extractor.cos_bucket_name, path=filename),
    )
    document_reference.set_client(extractor.watsonx_client)

    timestamp = int(time.time())
    cos_prefix = f"text_extraction_results/{pdf_path.stem}_{timestamp}/"
    results_reference = DataConnection(
        connection_asset_id=extractor.cos_connection_id,
        location=S3Location(bucket=extractor.cos_bucket_name, path=cos_prefix),
    )
    results_reference.set_client(extractor.watsonx_client)

    # Monkey-patch the extractor so it requests richer image handling without
    # mutating the upstream module's default behavior.
    extractor.create_parameters = _make_parameters_builder(embedded_images_mode)  # type: ignore[assignment]

    ok = extractor.run_multiple_formats_extraction(
        filename, document_reference, results_reference, cos_prefix
    )

    # Best-effort cleanup of the uploaded source file.
    try:
        extractor.cos_client.delete_object(
            Bucket=extractor.cos_bucket_name, Key=filename
        )
    except Exception:
        pass

    if not ok:
        return None

    return _download_cos_prefix(extractor, cos_prefix, local_output_dir)


def _download_cos_prefix(
    extractor: NotebookTextExtraction, cos_prefix: str, local_dir: Path
) -> Optional[Dict]:
    """Download every object under a COS prefix into local_dir and return
    info about the markdown + any asset files."""
    prefix = cos_prefix.rstrip("/")
    downloaded: List[str] = []
    markdown_text: Optional[str] = None
    markdown_key: Optional[str] = None

    # Paginate (watsonx can write hundreds of objects for image-rich docs).
    paginator = extractor.cos_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(
        Bucket=extractor.cos_bucket_name, Prefix=prefix
    ):
        for obj in page.get("Contents", []) or []:
            key = obj["Key"]
            rel = key[len(prefix):].lstrip("/")
            if not rel:
                continue
            out_path = local_dir / rel
            out_path.parent.mkdir(parents=True, exist_ok=True)
            extractor.cos_client.download_file(
                extractor.cos_bucket_name, key, str(out_path)
            )
            downloaded.append(str(out_path))

            if key.endswith("assembly.md") or (
                markdown_text is None and key.endswith(".md")
            ):
                with open(out_path, "r", encoding="utf-8") as fh:
                    markdown_text = fh.read()
                markdown_key = key

    if markdown_text is None:
        print("   ⚠️  No markdown file found in watsonx results")
        return None

    return {
        "markdown_text": markdown_text,
        "markdown_cos_key": markdown_key,
        "downloaded_files": downloaded,
    }


# --------------------------------------------------------------------------- #
# Main pipeline
# --------------------------------------------------------------------------- #

def _extract_single_xlsx(
    xlsx_path: Path,
    label: str,                     # user-visible tag, e.g. "sheet '품질지표'"
    per_doc_dir: Path,               # where images/pdf/watsonx subdirs live
    subdir_name: str,                # "" for whole-workbook; slug for per-sheet
    extractor: NotebookTextExtraction,
    skip_conversion: bool,
    libreoffice_cmd: Optional[str],
    embedded_images_mode: str,
) -> Optional[Dict]:
    """Run native-image extraction + LibreOffice → PDF + watsonx for one xlsx.

    Returns a dict with markdown_text / native_images / watsonx_assets / pdf_path,
    or None on failure. Paths are scoped under per_doc_dir[/<subdir_name>].
    """
    scope = per_doc_dir / subdir_name if subdir_name else per_doc_dir

    # 1. Native images (from the xlsx zip)
    native_images_dir = scope / "images"
    native_images = extract_images_from_workbook(xlsx_path, native_images_dir)
    print(f"     🖼️  {label}: {len(native_images)} native image(s)")

    # 2. xlsx → pdf
    pdf_dir = scope / "pdf"
    pdf_path: Optional[Path] = None
    existing_pdf = pdf_dir / f"{xlsx_path.stem}.pdf"
    if skip_conversion and existing_pdf.exists():
        pdf_path = existing_pdf
        print(f"     📄 {label}: reusing existing PDF {pdf_path.name}")
    else:
        if libreoffice_cmd is None:
            print(f"     ❌ {label}: LibreOffice not available")
            return None
        try:
            pdf_path = convert_excel_to_pdf(xlsx_path, pdf_dir, libreoffice_cmd)
            print(f"     ✅ {label}: PDF → {pdf_path.name}")
        except Exception as exc:
            print(f"     ❌ {label}: PDF conversion failed: {exc}")
            return None

    # 3. watsonx
    watsonx_dir = scope / "watsonx"
    try:
        result = extract_pdf_with_watsonx(
            extractor, pdf_path, watsonx_dir,
            embedded_images_mode=embedded_images_mode,
        )
    except Exception as exc:
        print(f"     ❌ {label}: watsonx extraction failed: {exc}")
        import traceback
        traceback.print_exc()
        return None

    if not result:
        return None

    md_text = result["markdown_text"]
    assets = [p for p in result["downloaded_files"] if not p.endswith(".md")]
    print(
        f"     ✅ {label}: watsonx md {len(md_text.encode('utf-8')):,} bytes, "
        f"{len(assets)} asset file(s)"
    )

    return {
        "pdf_path": str(pdf_path),
        "markdown_cos_key": result["markdown_cos_key"],
        "markdown_text": md_text,
        "native_images": native_images,
        "watsonx_assets": assets,
    }


def process_one_file(
    excel_path: Path,
    output_root: Path,
    extractor: NotebookTextExtraction,
    skip_conversion: bool,
    libreoffice_cmd: Optional[str],
    embedded_images_mode: str = _DEFAULT_EMBEDDED_IMAGES,
    per_sheet: bool = True,
) -> Optional[Dict]:
    stem = excel_path.stem
    per_doc_dir = output_root / stem
    per_doc_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*70}")
    print(f"📄 {excel_path.name}")
    print(f"   📂 Output: {per_doc_dir}")
    print(f"   📑 Mode:   {'per-sheet' if per_sheet else 'whole-workbook'}")
    print(f"{'='*70}")

    # ---- Whole-workbook mode: single extraction over the whole file. ----
    if not per_sheet or excel_path.suffix.lower() == ".xls":
        sub = _extract_single_xlsx(
            xlsx_path=excel_path,
            label="workbook",
            per_doc_dir=per_doc_dir,
            subdir_name="",
            extractor=extractor,
            skip_conversion=skip_conversion,
            libreoffice_cmd=libreoffice_cmd,
            embedded_images_mode=embedded_images_mode,
        )
        if not sub:
            return None
        return {
            "doc_id": str(uuid.uuid4()),
            "source_file": str(excel_path),
            "folder_name": stem,
            "output_dir": str(per_doc_dir),
            "mode": "whole_workbook",
            "pdf_path": sub["pdf_path"],
            "markdown_cos_key": sub["markdown_cos_key"],
            "extracted_text": sub["markdown_text"],
            "size_bytes": len(sub["markdown_text"].encode("utf-8")),
            "native_images": sub["native_images"],
            "watsonx_assets": sub["watsonx_assets"],
            "sheets": [],
        }

    # ---- Per-sheet mode: split, extract each, merge. ----
    split_dir = per_doc_dir / "sheets_split"
    try:
        splits = split_workbook_per_sheet(excel_path, split_dir)
    except Exception as exc:
        print(f"   ❌ Workbook split failed: {exc}")
        print("   ↪ Falling back to whole-workbook mode")
        return process_one_file(
            excel_path, output_root, extractor,
            skip_conversion, libreoffice_cmd,
            embedded_images_mode=embedded_images_mode,
            per_sheet=False,
        )
    print(f"   📑 Split into {len(splits)} sheet(s)")

    sheet_rows: List[Dict] = []
    merged_parts: List[str] = []
    all_native_images: List[str] = []
    all_assets: List[str] = []

    for sheet_idx, (sheet_name, sheet_xlsx) in enumerate(splits, 1):
        subdir = sheet_xlsx.stem  # e.g. "01_품질지표"
        label = f"[{sheet_idx}/{len(splits)}] sheet {sheet_name!r}"
        print(f"\n   ── {label} ──")

        sub = _extract_single_xlsx(
            xlsx_path=sheet_xlsx,
            label=label,
            per_doc_dir=per_doc_dir / "sheets",
            subdir_name=subdir,
            extractor=extractor,
            skip_conversion=skip_conversion,
            libreoffice_cmd=libreoffice_cmd,
            embedded_images_mode=embedded_images_mode,
        )

        if not sub:
            sheet_rows.append({
                "sheet_name": sheet_name,
                "sheet_index": sheet_idx,
                "status": "failed",
            })
            merged_parts.append(
                f"\n\n{'='*78}\n# Sheet {sheet_idx}: {sheet_name}\n"
                f"{'='*78}\n\n_(extraction failed)_\n"
            )
            continue

        sheet_rows.append({
            "sheet_name": sheet_name,
            "sheet_index": sheet_idx,
            "status": "ok",
            "split_xlsx": str(sheet_xlsx),
            "pdf_path": sub["pdf_path"],
            "markdown_cos_key": sub["markdown_cos_key"],
            "markdown_text": sub["markdown_text"],
            "size_bytes": len(sub["markdown_text"].encode("utf-8")),
            "native_images": sub["native_images"],
            "watsonx_assets": sub["watsonx_assets"],
        })
        all_native_images.extend(sub["native_images"])
        all_assets.extend(sub["watsonx_assets"])
        merged_parts.append(
            f"\n\n{'='*78}\n# Sheet {sheet_idx}: {sheet_name}\n"
            f"{'='*78}\n\n{sub['markdown_text'].rstrip()}\n"
        )

    merged_md = "".join(merged_parts).lstrip("\n")
    merged_md_path = per_doc_dir / "assembly.md"
    merged_md_path.write_text(merged_md, encoding="utf-8")
    print(f"\n   📝 Merged per-sheet markdown → {merged_md_path}")

    # Clean up the per-sheet split workbooks; we keep only `sheets/` (extracted
    # markdown + assets per sheet) and `assembly.md` as final outputs.
    if split_dir.exists():
        try:
            shutil.rmtree(split_dir)
        except OSError as exc:
            print(f"   ⚠️  Failed to clean up {split_dir}: {exc}")

    ok_count = sum(1 for r in sheet_rows if r["status"] == "ok")
    if ok_count == 0:
        return None

    return {
        "doc_id": str(uuid.uuid4()),
        "source_file": str(excel_path),
        "folder_name": stem,
        "output_dir": str(per_doc_dir),
        "mode": "per_sheet",
        "merged_markdown_path": str(merged_md_path),
        "extracted_text": merged_md,
        "size_bytes": len(merged_md.encode("utf-8")),
        "native_images": all_native_images,
        "watsonx_assets": all_assets,
        "sheets": sheet_rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Excel → PDF → watsonx Text Extraction → JSONL, with image capture."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="A single .xlsx/.xlsm/.xls file, or a folder to scan recursively.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("excel_extraction/output"),
        help="Root directory for per-document outputs (default: excel_extraction/output)",
    )
    parser.add_argument(
        "--jsonl",
        type=Path,
        default=None,
        help=(
            "Path to the aggregated JSONL file "
            "(default: <output-dir>/excel_extracted.jsonl)"
        ),
    )
    parser.add_argument(
        "--skip-conversion",
        action="store_true",
        help="Re-use existing PDFs under <output-dir>/<stem>/pdf/ instead of calling LibreOffice",
    )
    parser.add_argument(
        "--whole-workbook",
        action="store_true",
        help=(
            "Extract the workbook as a single document (no sheet separation). "
            "Default behavior splits each sheet into its own xlsx/PDF/watsonx "
            "run and merges results with per-sheet headers in assembly.md."
        ),
    )
    parser.add_argument(
        "--embedded-images",
        default=_DEFAULT_EMBEDDED_IMAGES,
        choices=[
            "disabled",
            "enabled_placeholder",
            "enabled_text",
            "enabled_verbalization",
            "enabled_verbalization_all",
        ],
        help=(
            "How watsonx handles images in the source PDF. "
            "'enabled_verbalization_all' (default) captions every image; "
            "'enabled_verbalization' only captions images likely to carry info; "
            "'enabled_text' OCRs text inside images; "
            "'enabled_placeholder' just inserts a placeholder string; "
            "'disabled' drops images entirely."
        ),
    )
    args = parser.parse_args()

    try:
        inputs = resolve_inputs(args.input_path)
    except (FileNotFoundError, ValueError) as exc:
        print(f"❌ {exc}")
        return 1

    if not inputs:
        print(f"❌ No Excel files found under {args.input_path}")
        return 1

    output_root = args.output_dir.resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    jsonl_path = (args.jsonl or output_root / "excel_extracted.jsonl").resolve()

    libreoffice_cmd = None if args.skip_conversion else check_libreoffice()
    if not args.skip_conversion and libreoffice_cmd is None:
        print("❌ LibreOffice not found. Install it or pass --skip-conversion.")
        print("   macOS:  brew install --cask libreoffice")
        print("   Linux:  sudo apt-get install libreoffice")
        return 1

    print("🚀 EXCEL → JSONL PIPELINE")
    print("=" * 70)
    print(f"   📥 Input:       {args.input_path} ({len(inputs)} file(s))")
    print(f"   📤 Output dir:  {output_root}")
    print(f"   📝 JSONL:       {jsonl_path}")
    print(f"   🔧 LibreOffice: {libreoffice_cmd or '(skipped)'}")
    print(f"   🖼️  Images:      {args.embedded_images}")
    print(f"   📑 Mode:        {'whole-workbook' if args.whole_workbook else 'per-sheet'}")
    print("=" * 70)

    print("\n🔧 Initializing watsonx extractor...")
    extractor = NotebookTextExtraction()

    results: List[Dict] = []
    failed: List[str] = []
    for idx, excel_path in enumerate(inputs, 1):
        print(f"\n[{idx}/{len(inputs)}]")
        try:
            row = process_one_file(
                excel_path, output_root, extractor,
                skip_conversion=args.skip_conversion,
                libreoffice_cmd=libreoffice_cmd,
                embedded_images_mode=args.embedded_images,
                per_sheet=not args.whole_workbook,
            )
        except Exception as exc:
            print(f"❌ Unhandled error on {excel_path.name}: {exc}")
            import traceback
            traceback.print_exc()
            row = None

        if row is None:
            failed.append(excel_path.name)
            continue

        results.append(row)
        # Append-as-we-go so partial runs still produce a usable JSONL.
        with open(jsonl_path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")

    print("\n" + "=" * 70)
    print("🎉 PIPELINE COMPLETE")
    print("=" * 70)
    print(f"   ✅ Success:  {len(results)}/{len(inputs)}")
    print(f"   ❌ Failed:   {len(failed)}/{len(inputs)}")
    if failed:
        for name in failed:
            print(f"      - {name}")
    print(f"   📝 JSONL:    {jsonl_path}")
    print(f"   📂 Per-doc:  {output_root}")
    return 0 if not failed else 2


if __name__ == "__main__":
    sys.exit(main())
